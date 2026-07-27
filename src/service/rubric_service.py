import hashlib
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Protocol

from pydantic import ValidationError

from schemas.rubric_contract import (
    CriterionEvaluation,
    RubricCreateRequest,
    RubricCriterion,
    RubricDefinition,
    RubricEvaluation,
    RubricEvaluationRequest,
    RubricScaleLevel,
)


class JSONChatClient(Protocol):
    def chat_json(
        self,
        system_prompt: str,
        user_prompt: str,
        temperature: float = 0.0,
    ) -> dict[str, Any]: ...


class RubricServiceError(ValueError):
    pass


class RubricStore:
    def __init__(self, root: str | Path = "data/rubrics") -> None:
        self.root = Path(root)
        self.root.mkdir(parents=True, exist_ok=True)

    def save(self, rubric: RubricDefinition) -> Path:
        path = self.root / f"{rubric.rubric_id}.json"
        path.write_text(
            rubric.model_dump_json(indent=2) + "\n",
            encoding="utf-8",
        )
        return path

    def get(self, rubric_id: str) -> RubricDefinition:
        path = self.root / f"{rubric_id}.json"
        if not path.is_file():
            raise FileNotFoundError(f"Rubric not found: {rubric_id}")
        return RubricDefinition.model_validate_json(
            path.read_text(encoding="utf-8")
        )

    def list_for_session(
        self, course_id: str, class_session_id: str
    ) -> list[RubricDefinition]:
        results: list[RubricDefinition] = []
        for path in self.root.glob("*.json"):
            try:
                rubric = RubricDefinition.model_validate_json(
                    path.read_text(encoding="utf-8")
                )
            except Exception:
                continue
            if (
                rubric.course_id == course_id
                and rubric.class_session_id == class_session_id
            ):
                results.append(rubric)
        return sorted(results, key=lambda item: item.created_at)


class RubricService:
    def __init__(
        self,
        store: RubricStore | None = None,
        llm_client: JSONChatClient | None = None,
    ) -> None:
        self.store = store or RubricStore()
        self.llm_client = llm_client

    def create(self, request: RubricCreateRequest) -> RubricDefinition:
        self._validate_weights(request.criteria)
        digest = hashlib.sha256(
            request.model_dump_json().encode("utf-8")
        ).hexdigest()[:16]
        rubric = RubricDefinition(
            **request.model_dump(),
            rubric_id=f"rubric-{digest}",
            created_at=datetime.now(timezone.utc),
        )
        self.store.save(rubric)
        return rubric

    def create_from_file(
        self,
        *,
        course_id: str,
        class_session_id: str,
        title: str,
        filename: str,
        content: bytes,
    ) -> RubricDefinition:
        text = self._extract_text(filename, content)
        if not text.strip():
            raise RubricServiceError("Rubric file contains no readable text")

        if self.llm_client is not None:
            payload = self.llm_client.chat_json(
                system_prompt=self._EXTRACTION_PROMPT,
                user_prompt=json.dumps(
                    {
                        "course_id": course_id,
                        "class_session_id": class_session_id,
                        "title": title,
                        "rubric_text": text[:30000],
                    },
                    ensure_ascii=False,
                ),
            )
            try:
                request = RubricCreateRequest.model_validate(payload)
            except ValidationError as exc:
                raise RubricServiceError(
                    "LLM returned an invalid rubric structure"
                ) from exc
        else:
            request = self._fallback_rubric(
                course_id=course_id,
                class_session_id=class_session_id,
                title=title,
                text=text,
            )

        rubric = self.create(request)
        rubric = rubric.model_copy(update={"source_filename": filename})
        self.store.save(rubric)
        return rubric

    def evaluate(
        self,
        rubric_id: str,
        request: RubricEvaluationRequest,
    ) -> RubricEvaluation:
        rubric = self.store.get(rubric_id)
        if self.llm_client is not None:
            raw = self.llm_client.chat_json(
                system_prompt=self._EVALUATION_PROMPT,
                user_prompt=json.dumps(
                    {
                        "rubric": rubric.model_dump(mode="json"),
                        "submission": request.model_dump(mode="json"),
                    },
                    ensure_ascii=False,
                ),
            )
            return self._validate_llm_evaluation(
                raw=raw,
                rubric=rubric,
                request=request,
            )
        return self._deterministic_evaluation(rubric, request)

    _EXTRACTION_PROMPT = """
Convert the supplied rubric document into exactly one JSON object matching:
{"course_id":"string","class_session_id":"string","title":"string","description":"string","criteria":[{"criterion_id":"string","title":"string","description":"string","weight":0.5,"levels":[{"label":"string","score":4,"description":"string"}]}]}
Weights must sum to 1. Preserve the source meaning. No markdown.
""".strip()

    _EVALUATION_PROMPT = """
You are a strict but constructive assessment assistant. Evaluate only against the supplied rubric. Return exactly one JSON object:
{"criterion_results":[{"criterion_id":"string","score":0,"max_score":4,"feedback":"string","evidence":["string"]}],"overall_feedback":"string","strengths":["string"],"improvements":["string"]}
Every rubric criterion must appear exactly once. Scores must be within each criterion's scale. Do not invent evidence.
""".strip()

    @staticmethod
    def _validate_weights(criteria: list[RubricCriterion]) -> None:
        total = sum(item.weight for item in criteria)
        if abs(total - 1.0) > 0.001:
            raise RubricServiceError("Rubric criterion weights must sum to 1")

    def _validate_llm_evaluation(
        self,
        *,
        raw: dict[str, Any],
        rubric: RubricDefinition,
        request: RubricEvaluationRequest,
    ) -> RubricEvaluation:
        expected = {item.criterion_id: item for item in rubric.criteria}
        raw_results = raw.get("criterion_results")
        if not isinstance(raw_results, list):
            raise RubricServiceError("LLM evaluation is missing criterion_results")

        results: list[CriterionEvaluation] = []
        for item in raw_results:
            criterion_id = str(item.get("criterion_id", ""))
            criterion = expected.get(criterion_id)
            if criterion is None:
                raise RubricServiceError("LLM returned an unknown criterion")
            max_score = max(level.score for level in criterion.levels)
            score = min(max(float(item.get("score", 0)), 0), max_score)
            results.append(
                CriterionEvaluation(
                    criterion_id=criterion_id,
                    score=score,
                    max_score=max_score,
                    feedback=str(item.get("feedback", "No feedback provided")),
                    evidence=[str(value) for value in item.get("evidence", [])],
                )
            )

        if {item.criterion_id for item in results} != set(expected):
            raise RubricServiceError("LLM did not evaluate every criterion")
        return self._build_evaluation(rubric, request, results, raw)

    def _deterministic_evaluation(
        self,
        rubric: RubricDefinition,
        request: RubricEvaluationRequest,
    ) -> RubricEvaluation:
        submission_tokens = set(re.findall(r"[\w-]+", request.submission_text.lower()))
        results: list[CriterionEvaluation] = []
        for criterion in rubric.criteria:
            criterion_tokens = set(
                re.findall(
                    r"[\w-]+",
                    f"{criterion.title} {criterion.description}".lower(),
                )
            )
            overlap = len(submission_tokens & criterion_tokens)
            ratio = min(overlap / max(len(criterion_tokens), 1), 1.0)
            max_score = max(level.score for level in criterion.levels)
            score = round(max_score * ratio, 2)
            results.append(
                CriterionEvaluation(
                    criterion_id=criterion.criterion_id,
                    score=score,
                    max_score=max_score,
                    feedback=(
                        "Evidence related to this criterion was found."
                        if overlap
                        else "No clear evidence for this criterion was found."
                    ),
                    evidence=request.evidence[:3],
                )
            )
        return self._build_evaluation(rubric, request, results, {})

    @staticmethod
    def _build_evaluation(
        rubric: RubricDefinition,
        request: RubricEvaluationRequest,
        results: list[CriterionEvaluation],
        raw: dict[str, Any],
    ) -> RubricEvaluation:
        by_id = {item.criterion_id: item for item in rubric.criteria}
        weighted_total = 0.0
        max_total = 100.0
        for result in results:
            criterion = by_id[result.criterion_id]
            weighted_total += (
                result.score / result.max_score
            ) * criterion.weight * 100
        now = datetime.now(timezone.utc)
        digest = hashlib.sha256(
            f"{rubric.rubric_id}:{request.student_id}:{now.isoformat()}".encode()
        ).hexdigest()[:16]
        return RubricEvaluation(
            evaluation_id=f"evaluation-{digest}",
            rubric_id=rubric.rubric_id,
            student_id=request.student_id,
            total_score=round(weighted_total, 2),
            max_score=max_total,
            percentage=round(weighted_total, 2),
            criterion_results=results,
            overall_feedback=str(raw.get("overall_feedback", "Evaluation completed.")),
            strengths=[str(value) for value in raw.get("strengths", [])],
            improvements=[str(value) for value in raw.get("improvements", [])],
            created_at=now,
        )

    @staticmethod
    def _fallback_rubric(
        *, course_id: str, class_session_id: str, title: str, text: str
    ) -> RubricCreateRequest:
        lines = [line.strip(" -•\t") for line in text.splitlines() if len(line.strip()) > 3]
        lines = lines[:10] or ["Overall quality"]
        weight = 1 / len(lines)
        criteria = []
        for index, line in enumerate(lines, start=1):
            criteria.append(
                RubricCriterion(
                    criterion_id=f"criterion-{index}",
                    title=line[:120],
                    description=line,
                    weight=weight,
                    levels=[
                        RubricScaleLevel(label="Needs improvement", score=1, description="Limited evidence"),
                        RubricScaleLevel(label="Developing", score=2, description="Partial evidence"),
                        RubricScaleLevel(label="Proficient", score=3, description="Clear evidence"),
                        RubricScaleLevel(label="Excellent", score=4, description="Complete and accurate evidence"),
                    ],
                )
            )
        criteria[-1] = criteria[-1].model_copy(
            update={"weight": 1 - sum(item.weight for item in criteria[:-1])}
        )
        return RubricCreateRequest(
            course_id=course_id,
            class_session_id=class_session_id,
            title=title,
            description="Imported from uploaded rubric file",
            criteria=criteria,
        )

    @staticmethod
    def _extract_text(filename: str, content: bytes) -> str:
        suffix = Path(filename).suffix.lower()
        if suffix in {".txt", ".md", ".csv"}:
            return content.decode("utf-8", errors="replace")
        if suffix == ".pdf":
            import fitz
            document = fitz.open(stream=content, filetype="pdf")
            try:
                return "\n".join(page.get_text("text") for page in document)
            finally:
                document.close()
        if suffix == ".docx":
            from io import BytesIO
            from docx import Document
            document = Document(BytesIO(content))
            return "\n".join(paragraph.text for paragraph in document.paragraphs)
        if suffix == ".xlsx":
            from io import BytesIO
            from openpyxl import load_workbook
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
            rows: list[str] = []
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    rows.append(" | ".join("" if value is None else str(value) for value in row))
            return "\n".join(rows)
        raise RubricServiceError("Supported rubric files: PDF, DOCX, XLSX, TXT, MD, CSV")
